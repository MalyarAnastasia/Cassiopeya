#!/usr/bin/env python3
"""
Конвертер CSV в XLSX для Pascal Legacy
Обрабатывает правильные типы данных: timestamp, boolean, числа, строки
"""

import csv
import os
import sys
from datetime import datetime
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


def parse_boolean(value):
    """Парсит boolean значения: ИСТИНА/ЛОЖЬ, true/false, 1/0"""
    val = str(value).strip().upper()
    if val in ('ИСТИНА', 'TRUE', '1', 'YES', 'ДА'):
        return True
    elif val in ('ЛОЖЬ', 'FALSE', '0', 'NO', 'НЕТ'):
        return False
    return None


def parse_timestamp(value):
    """Парсит timestamp в различных форматах"""
    if not value:
        return None
    
    # ISO 8601 формат
    formats = [
        '%Y-%m-%dT%H:%M:%SZ',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    
    return value  # Возвращаем как строку если не распарсилось


def convert_csv_to_xlsx(csv_path, xlsx_path=None):
    """Конвертирует CSV в XLSX с правильными типами данных"""
    csv_path = Path(csv_path)
    
    if not csv_path.exists():
        print(f"Error: CSV file not found: {csv_path}")
        return False
    
    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetry Data"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        
        # Записываем заголовки
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Записываем данные с правильными типами
        for row_idx, row in enumerate(reader, 2):
            for col_idx, value in enumerate(row, 1):
                header = headers[col_idx - 1].lower()
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Определяем тип данных по заголовку
                if 'timestamp' in header or 'recorded_at' in header or 'date' in header:
                    dt = parse_timestamp(value)
                    if isinstance(dt, datetime):
                        cell.value = dt
                        cell.number_format = 'yyyy-mm-dd hh:mm:ss'
                    else:
                        cell.value = value
                
                elif 'bool' in header or 'is_' in header or 'active' in header:
                    bool_val = parse_boolean(value)
                    if bool_val is not None:
                        cell.value = bool_val
                    else:
                        cell.value = value
                
                elif 'voltage' in header or 'temp' in header or 'value' in header:
                    try:
                        cell.value = float(value)
                        cell.number_format = '0.00'
                    except ValueError:
                        cell.value = value
                
                else:
                    cell.value = value
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Автоматическая ширина столбцов
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                if row[0].value:
                    max_length = max(max_length, len(str(row[0].value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    wb.save(xlsx_path)
    print(f"Successfully converted {csv_path} to {xlsx_path}")
    return True


def process_directory(csv_dir):
    """Обрабатывает все CSV файлы в директории"""
    csv_dir = Path(csv_dir)
    csv_files = list(csv_dir.glob('telemetry_*.csv'))
    
    if not csv_files:
        print(f"No CSV files found in {csv_dir}")
        return
    
    for csv_file in csv_files:
        xlsx_file = csv_file.with_suffix('.xlsx')
        if not xlsx_file.exists() or csv_file.stat().st_mtime > xlsx_file.stat().st_mtime:
            convert_csv_to_xlsx(csv_file, xlsx_file)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
        if os.path.isdir(csv_path):
            process_directory(csv_path)
        else:
            convert_csv_to_xlsx(csv_path)
    else:
        csv_dir = os.getenv('CSV_OUT_DIR', '/data/csv')
        process_directory(csv_dir)



